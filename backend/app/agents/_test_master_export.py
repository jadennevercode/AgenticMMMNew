"""2.6's model-input export: one indicator set across every surface, response included.

The granularity reference, the Data Station and the factor-tree close-out all
answer "which indicators is the model built on". They used to answer it three
different ways and disagree — 7 / 5 / 8 on the drill case — because:

* the **response has no ledger row** (no layer rules on the thing the drivers
  explain), so the two surfaces that filtered on the ledger dropped it, and the
  exported 2.32 "model input" shipped with no dependent variable at all;
* the **granularity sheet joined in the wrong key space**, matching the factor
  row's declared ``(l4, indicator)`` against the data's ``(l4, metric)`` — which
  differ precisely when a mapping exists (see `app.agents.factor_link`).

All three now derive from `master_data.adopted_indicators`.

Run: PYTHONPATH=. .venv/bin/python -m app.agents._test_master_export
"""
from __future__ import annotations

import io
import sys

from app.agents import master_data as md
from app.agents._test_per_channel import make_two_channel_state
from app.agents.master_data import build_export


# ── pure units ──────────────────────────────────────────────────────────────

def test_sheet_titles_are_excel_legal_and_unique() -> None:
    used: set[str] = set()
    assert md._safe_sheet_title("MT · B", used) == "MT · B"
    # openpyxl raises on a duplicate title, so collisions must be resolved here.
    assert md._safe_sheet_title("MT · B", used) == "MT · B~2"
    t = md._safe_sheet_title("a[b]c:d*e?f/g\\h" + "x" * 40, used)
    assert len(t) <= 31 and not set(t) & set("[]:*?/\\"), t
    assert md._safe_sheet_title("", used) == "model"


def test_scope_strings_follow_the_2_32_convention() -> None:
    assert md._channel_scope([]) == "全渠道"          # national / all-channel data
    assert md._channel_scope(["MT", "EC"]) == "MT,EC"
    assert md._region_scope(["National"]) == "National"
    assert md._region_scope(["B", "A"]) == "A,B"


def test_text_does_not_render_pandas_missing_as_a_factor_level() -> None:
    """`astype("string")` yields the literal "<NA>"; printed into a factor sheet
    it reads as a real level."""
    assert md._text("<NA>") == "" and md._text("nan") == "" and md._text(None) == ""
    assert md._text("  KPI ") == "KPI"


# ── against the two-channel fixture ─────────────────────────────────────────

def test_export_has_a_sheet_per_model_object() -> None:
    st = make_two_channel_state("t-master-export")
    data = build_export(st)
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # Sheets are named by model object — a (channel × product) cell, not a bare
    # channel type, since 2026-07-27.
    assert wb.sheetnames[:2] == ["模型颗粒度参考表", "D.Data Station"], wb.sheetnames
    assert {"MT · B", "TT · B"} <= set(wb.sheetnames), wb.sheetnames
    ws = wb["MT · B"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header and header[0] == "Period", header
    assert len(header) >= 2, header
    print(f"  export ok: sheets={wb.sheetnames}, MT cols={len(header)}")


def test_the_response_reaches_the_export() -> None:
    """The bug this file exists for: the whole of Y was filtered out of the
    model-input deliverable, because Y is not a driver and has no ledger row."""
    st = make_two_channel_state("t-master-response")
    adopted = md.adopted_indicators(st)
    responses = {k: v for k, v in adopted.items() if v["role"] == "response"}
    assert responses, "the model input has no dependent variable"

    ds = md.data_station(st, limit=10_000_000)
    l4_col = ds["columns"].index("数据类型Level4")
    metric_col = ds["columns"].index("METRICS")
    shipped = {(str(r[l4_col]).strip().lower(), str(r[metric_col]).strip().lower())
               for r in ds["rows"]}
    for key in responses:
        assert key in shipped, f"the response {key} is missing from the Data Station"


def test_the_granularity_sheet_and_the_data_station_agree() -> None:
    st = make_two_channel_state("t-master-agree")
    adopted = md.adopted_indicators(st)
    gref = [r for r in md.granularity_reference(st) if r["adopted"]]
    assert len(gref) == len(adopted), (
        f"granularity={len(gref)} adopted={len(adopted)} — the export must not "
        "describe a different indicator set than the model was built on")

    ds = md.data_station(st, limit=10_000_000)
    l4_col = ds["columns"].index("数据类型Level4")
    metric_col = ds["columns"].index("METRICS")
    shipped = {(str(r[l4_col]).strip().lower(), str(r[metric_col]).strip().lower())
               for r in ds["rows"]}
    assert shipped <= set(adopted), f"the Data Station carries un-adopted rows: {shipped - set(adopted)}"


def test_the_data_station_export_is_uncapped() -> None:
    """The UI pages the station; the deliverable must not."""
    st = make_two_channel_state("t-master-uncapped")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(build_export(st)))
    assert wb["D.Data Station"].max_row - 1 == md.data_station(st, limit=10_000_000)["rowCount"]


def test_a_national_row_follows_its_own_model_not_the_union() -> None:
    """A row with no channel is national — shared into *every* model — so one
    channel rejecting it must not delete it from the channel that kept it.

    Screening those rows against the union of every object's excludes reads as the
    safe direction and is not: on the drill case TT was fitted with 温度 at 43.3%
    contribution while the master table deleted it from TT's own wide table,
    because MT and EC had rejected it. The deliverable then described a model
    nobody fitted.
    """
    from app.agents.dataset_cache import model_df, model_objects
    from app.agents.ledger import OBJECT_ANY
    from app.domain.models import OlsRangeRow, OlsRangeScorecard

    st = make_two_channel_state("t-master-national")
    df = model_df(st)
    # The fixture's rows all carry a channel, so make one national by hand: this
    # is about the mask's treatment of an unpinned row, not about the fixture.
    national = df["metric"] == "广告投放"
    df.loc[national, "channel_type"] = None
    df.loc[national, "channel"] = None
    from app.agents.dataset_cache import invalidate_project, set_project_dataset
    invalidate_project(st.project_id)
    set_project_dataset(st.project_id, df, "slot")

    objects = model_objects(st)
    assert len(objects) >= 2, objects
    rejecting, keeping = objects[0], objects[1]
    st.ols_scorecard = OlsRangeScorecard(rows=[OlsRangeRow(
        id=f"{rejecting}|广告投放|广告投放", object=rejecting,
        l4="广告投放", indicator="广告投放", metric="广告投放",
        autoVerdict="reject", disposition="reject", decidedBy="human")])

    kept = md._adopted_df(st, scope=[keeping])
    assert "广告投放" in set(kept["metric"].astype(str)), (
        f"{rejecting} rejected it; {keeping} kept it and must still carry it")
    dropped = md._adopted_df(st, scope=[rejecting])
    assert "广告投放" not in set(dropped["metric"].astype(str)), (
        f"{rejecting} rejected it and must not carry it")


def test_each_sheet_carries_exactly_its_own_model_input() -> None:
    """The per-object sheet must use the row selection the FIT uses. Slicing by
    brand + channel_type looks equivalent and drops every national row and every
    competitor row — the shared rows the model is actually fitted on."""
    st = make_two_channel_state("t-master-sheets")
    from app.agents.dataset_cache import model_objects
    from app.agents.model_objects import object_label

    sheets = {label: cols for label, cols, _rows in md.model_input_sheets(st)}
    for obj in model_objects(st):
        label = object_label(obj)
        assert label in sheets, f"{label} has no sheet"
        assert sheets[label][0] == "Period"
        assert len(sheets[label]) >= 2, f"{label} carries no indicator at all"


def test_a_rejected_indicator_leaves_every_surface() -> None:
    """The point of deriving all three from one place: a verdict has to reach the
    deliverable, not just the scorecard it was recorded on."""
    st = make_two_channel_state("t-master-reject")
    before = md.adopted_indicators(st)
    victim = next((k for k, v in before.items() if v["role"] == "driver"), None)
    assert victim is not None, "fixture has no driver to reject"

    from app.agents.ledger import OBJECT_ANY
    from app.domain.models import OlsRangeRow, OlsRangeScorecard
    st.ols_scorecard = OlsRangeScorecard(rows=[OlsRangeRow(
        id=f"{OBJECT_ANY}|{victim[0]}|{victim[1]}", object=OBJECT_ANY,
        l4=victim[0], indicator=victim[1], metric=victim[1],
        autoVerdict="reject", disposition="reject", decidedBy="human")])

    after = md.adopted_indicators(st)
    assert victim not in after, "a rejected indicator must leave the adopted set"
    ds = md.data_station(st, limit=10_000_000)
    l4_col = ds["columns"].index("数据类型Level4")
    metric_col = ds["columns"].index("METRICS")
    shipped = {(str(r[l4_col]).strip().lower(), str(r[metric_col]).strip().lower())
               for r in ds["rows"]}
    assert victim not in shipped, "a rejected indicator must leave the Data Station"


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failed = 0
    for fn in fns:
        try:
            fn()
            print(f"  ✓ {fn.__name__}")
        except Exception as e:  # noqa: BLE001
            failed += 1
            print(f"  FAIL {fn.__name__}: {e}")
    print(f"{len(fns) - failed}/{len(fns)} passed")
    sys.exit(1 if failed else 0)
