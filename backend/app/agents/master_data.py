"""2.6 Master Data — the modeling feature table, sliced on the 2.24 dimensions.

The modeling long table (the reference ``Data Process_2.24.xlsx`` schema) carries
per row: brand / province_group / channel_type / channel · year / month · the
L1–L8 factor path · METRICS type + label · VALUE. The master table is that long
table

  * restricted to the indicators the **ledger** reports as adopted — an
    indicator any S2 layer rejected can never appear here, and
  * pivoted to one column per indicator over the chosen time grain.

It is computed **live** rather than baked into the artifact: the user slices by
product × channel × region, and materializing every combination up front would
be both enormous and stale the moment a verdict changes.
"""
from __future__ import annotations

from typing import Optional

import pandas as pd

from app.agents.ledger import model_selection
from app.dataeng.validation_query import (
    _available_grains,
    _kpi_mask,
    _period_keys,
    _period_label,
)
from app.store.state import ProjectState

# filter key (camelCase, as the API takes it) → long-table column. `channel` is
# in here and not in the validation view's set: the master table is what a user
# slices per-channel before handing it to modeling.
DIM_COLS: dict[str, str] = {
    "brand": "brand",
    "provinceGroup": "province_group",
    "channelType": "channel_type",
    "channel": "channel",
}

# A wide table is for reading, not for scrolling forever.
MAX_ROWS = 400
MAX_COLS = 60


def _lower(s: object) -> str:
    return str(s).strip().lower() if s is not None else ""


def _text(s: object) -> str:
    """A display string from a pandas cell — `astype("string")` turns NaN into the
    literal "<NA>", which would otherwise be rendered as a factor level."""
    v = str(s).strip() if s is not None else ""
    return "" if v in ("<NA>", "nan", "None") else v


def _distinct(df: pd.DataFrame, col: str) -> list[str]:
    if col not in df.columns:
        return []
    vals = df[col].astype("string").str.strip().dropna().unique().tolist()
    return sorted(v for v in vals if v and v.lower() not in ("nan", "none", "na"))


def _apply_dims(df: pd.DataFrame, dims: dict[str, Optional[list[str]]]) -> pd.DataFrame:
    out = df
    for key, col in DIM_COLS.items():
        wanted = dims.get(key)
        if wanted and col in out.columns:
            keep = {_lower(v) for v in wanted}
            out = out[out[col].astype("string").str.strip().str.lower().isin(keep)]
    return out


def adopted_mask(st: ProjectState, df: pd.DataFrame,
                 *, scope: Optional[list[str]] = None) -> pd.Series:
    """Rows whose indicator survived every S2 filter layer (plus the KPI rows).

    The KPI is the response, not a factor — no layer rules on it, so it is never
    in the ledger and must be admitted explicitly or the table loses its Y.

    ``model_selection``'s ``exclude``/``include`` are keyed per model object
    (channel type) so 2.5r/2.6/3.2's per-object fits never let one channel's
    drop or tick leak into another's. This mask resolves the keep/drop
    decision **per channel_type present in df** — a master-table call is
    already sliced to (at most) a handful of channels via ``_apply_dims``, so
    it walks each one's own rows against its own ``exclude_for``/
    ``include_for`` rather than flattening every object's verdicts into one
    union. That is the whole point of per-channel screening: a wide table
    sliced to one Channel Type shows exactly that channel's surviving
    indicators, and the same indicator can appear as a column in one channel
    and be absent in another.

    A row that cannot be pinned to one model object — blank ``channel_type``, or a
    ``brand`` no model object claims — is **national**: it is shared into every
    object's fit rather than belonging to one. It is therefore kept when any object
    in ``scope`` kept it, and dropped only when every one of them rejected it.
    ``scope`` defaults to every model object; pass a single object to get exactly
    that model's own input (what the per-object export sheets do). Note
    ``.astype("string")`` turns NaN into the
    literal ``"<NA>"``, which left unnormalized the per-object loop would treat as
    a real object — ``exclude_for("<NA>")`` resolves only the OBJECT_ANY-level
    drops, so a channel-specific rejection could leak in through any unmapped row.
    Hence the normalization to ``""`` up front.

    The object key is ``(channel_type, brand)`` since 2026-07-27, so the loop walks
    the composite ids ``model_selection`` is keyed by. Keying it on the bare
    channel type would have silently resolved every row through the OBJECT_ANY
    bucket and let each channel's own rejections through.
    """
    from app.agents.dataset_cache import model_objects
    from app.agents.model_objects import make_object

    sel = model_selection(st)
    l4 = df["l4"].astype("string").map(_lower) if "l4" in df.columns else pd.Series("", index=df.index)
    metric = df["metric"].astype("string").map(_lower)
    if "channel_type" in df.columns:
        raw_ct = df["channel_type"].astype("string")
        ct_only = raw_ct.map(lambda s: "" if pd.isna(s) or str(s).strip() in ("", "<NA>") else str(s).strip())
    else:
        ct_only = pd.Series("", index=df.index)
    if "brand" in df.columns:
        raw_b = df["brand"].astype("string")
        brand = raw_b.map(lambda s: "" if pd.isna(s) or str(s).strip() in ("", "<NA>") else str(s).strip())
    else:
        brand = pd.Series("", index=df.index)
    known = set(model_objects(st))
    # A row belongs to its (channel, product) object when that object exists;
    # otherwise it is unpinned and takes the strict union path below.
    ct = pd.Series([make_object(c, b) if make_object(c, b) in known else ""
                    for c, b in zip(ct_only, brand)], index=df.index)
    keep = pd.Series(True, index=df.index)
    for obj in sorted(set(ct) - {""}):
        excl = sel.exclude_for(obj)
        metric_only = {m for excl_l4, m in excl if not excl_l4}
        inc = sel.include_for(obj)
        rowsel = ct == obj
        rej = pd.Series([(a, b) in excl or b in metric_only
                         for a, b in zip(l4[rowsel], metric[rowsel])], index=l4[rowsel].index)
        if inc is None:
            block = rej
        else:
            # `inc` is keyed (norm_l4, norm_metric); legacy configs may still
            # carry bare metric names, which keep that metric under any L4.
            inc_pairs = {i for i in inc if isinstance(i, tuple)}
            inc_metrics = {i for i in inc if isinstance(i, str)}
            picked = pd.Series([(a, b) in inc_pairs or b in inc_metrics
                                for a, b in zip(l4[rowsel], metric[rowsel])],
                               index=l4[rowsel].index)
            block = rej | ~picked
        keep.loc[rowsel] = ~block

    unmapped = ct == ""
    if unmapped.any():
        # A row with no channel is **national** — media bought once for the country,
        # shared into every model object (the 2026-07-27 rule). So it is part of
        # model X exactly when X's own screening kept it, and it belongs in this
        # table when *any* object in scope kept it.
        #
        # This used to drop it when any object excluded it (the union of every
        # object's excludes). That reads as the safe direction and is not: on the
        # drill case the TT model was fitted with 温度 at 43.3% contribution while
        # this mask deleted it from TT's own wide table, because MT and EC had
        # rejected it. The deliverable then described a model that was never fitted
        # — the exact failure this table exists to prevent.
        objects = scope if scope is not None else sorted(known)
        if not objects:
            rej = pd.Series([_matches_excl((a, b), sel.exclude_for(""))
                             for a, b in zip(l4[unmapped], metric[unmapped])],
                            index=l4[unmapped].index)
        else:
            rej = pd.Series(
                [all(_matches_excl((a, b), sel.exclude_for(o)) for o in objects)
                 for a, b in zip(l4[unmapped], metric[unmapped])],
                index=l4[unmapped].index)
        keep.loc[unmapped] = keep.loc[unmapped] & ~rej

    return keep | _kpi_mask(df)


def _matches_excl(key: tuple[str, str], excl: frozenset[tuple[str, str]]) -> bool:
    """Mirror `build_model_frame`'s exclude semantics: an exact (l4, metric) hit, or
    a metric-only entry (empty l4) dropping that metric under any L4."""
    return key in excl or any(not l4 and m == key[1] for l4, m in excl)


def _adopted_df(st: ProjectState, *, scope: Optional[list[str]] = None) -> pd.DataFrame:
    """The modeling rows that survived every S2 layer.

    ``scope`` narrows which model objects a **national** row (no channel of its
    own) is screened against — pass one object to get exactly that model's input.
    """
    from app.agents.dataset_cache import model_df
    df = model_df(st)
    return df[adopted_mask(st, df, scope=scope)]


def dimensions(st: ProjectState) -> dict:
    """The slicing options the master table offers, from the adopted rows only."""
    try:
        df = _adopted_df(st)
    except Exception:  # noqa: BLE001 — no bound data yet
        return {"brand": [], "provinceGroup": [], "channelType": [], "channel": [],
                "grains": ["month"], "indicators": []}
    return {
        "brand": _distinct(df, "brand"),
        "provinceGroup": _distinct(df, "province_group"),
        "channelType": _distinct(df, "channel_type"),
        "channel": _distinct(df, "channel"),
        "grains": _available_grains(df),
        "indicators": _distinct(df, "metric"),
    }


def master_table(
    st: ProjectState,
    *,
    brand: Optional[list[str]] = None,
    province_group: Optional[list[str]] = None,
    channel_type: Optional[list[str]] = None,
    channel: Optional[list[str]] = None,
    grain: str = "month",
    indicators: Optional[list[str]] = None,
) -> dict:
    """The adopted feature wide table for one product × channel × region slice.

    Returns ``{columns, rows, kpi, truncated, rowCount, colCount, note}`` where
    each row is one period and each column one adopted indicator.
    """
    try:
        df = _adopted_df(st)
    except Exception as e:  # noqa: BLE001
        return {"columns": [], "rows": [], "kpi": "", "truncated": False,
                "rowCount": 0, "colCount": 0, "note": f"No modeling data available: {e}"}

    df = _apply_dims(df, {"brand": brand, "provinceGroup": province_group,
                          "channelType": channel_type, "channel": channel})
    if df.empty:
        return {"columns": [], "rows": [], "kpi": "", "truncated": False,
                "rowCount": 0, "colCount": 0,
                "note": "No rows match this slice — widen the filters."}

    if grain not in _available_grains(df):
        grain = "month" if "month" in _available_grains(df) else "year"

    # The primary KPI (the response the wide table explains) is the response the
    # user set at 2.1 and the model was fitted on — resolved through the one
    # authority so the master table cannot headline a different Y than the fit.
    # Falls back to the Volume-preferring auto-pick for un-configured projects.
    kpi_rows = df[_kpi_mask(df)]
    kpi_metric = ""
    if not kpi_rows.empty:
        from app.agents.indicator_metadata import classify_indicator
        from app.agents.overrides import resolved_y_metric
        kpi_names = _distinct(kpi_rows, "metric")
        resolved = resolved_y_metric(st, df) or ""
        if resolved and any(_lower(resolved) == _lower(m) for m in kpi_names):
            kpi_metric = next(m for m in kpi_names if _lower(m) == _lower(resolved))
        else:
            volume = [m for m in kpi_names if classify_indicator(m).metric_type == "kpi_volume"]
            kpi_metric = volume[0] if volume else (
                str(kpi_rows["metric"].mode().iloc[0]) if not kpi_rows["metric"].mode().empty else "")

    if indicators:
        wanted = {_lower(i) for i in indicators} | {_lower(kpi_metric)}
        df = df[df["metric"].astype("string").str.strip().str.lower().isin(wanted)]

    keys = _period_keys(df, grain)
    frame = pd.DataFrame({
        "_k": keys,
        "_m": df["metric"].astype("string").str.strip(),
        "_v": pd.to_numeric(df["value"], errors="coerce"),
    }).dropna(subset=["_k", "_m"])
    if frame.empty:
        return {"columns": [], "rows": [], "kpi": kpi_metric, "truncated": False,
                "rowCount": 0, "colCount": 0, "note": "No usable values in this slice."}

    # DATA-011/007: roll each indicator up to its period value with ITS OWN aggregation
    # — a rate/coverage metric (NDWD) averaged across the sliced dimensions, spend and
    # volume summed. A single pivot aggfunc would wrongly sum a rate over regions.
    from app.dataeng.validation_query import _metric_agg, _pandas_agg
    cols_data: dict[str, pd.Series] = {}
    for m, g in frame.groupby("_m"):
        cols_data[str(m)] = g.groupby("_k")["_v"].agg(_pandas_agg(_metric_agg(st, str(m))))
    wide = pd.DataFrame(cols_data).sort_index()

    # KPI first — it is the response the rest of the table explains.
    cols = [c for c in wide.columns if _lower(c) == _lower(kpi_metric)]
    cols += sorted(c for c in wide.columns if _lower(c) != _lower(kpi_metric))
    truncated = len(cols) > MAX_COLS or len(wide.index) > MAX_ROWS
    cols = cols[:MAX_COLS]
    index = list(wide.index)[-MAX_ROWS:]

    rows = [[_period_label(int(k), grain)]
            + [None if pd.isna(v := wide.at[k, c]) else round(float(v), 2) for c in cols]
            for k in index]
    return {
        "columns": ["Period"] + [str(c) for c in cols],
        "rows": rows,
        "kpi": kpi_metric,
        "grain": grain,
        "truncated": truncated,
        "rowCount": len(index),
        "colCount": len(cols),
        "note": (f"Showing the last {MAX_ROWS} periods / first {MAX_COLS} indicators."
                 if truncated else ""),
    }


# ── 2.32 reference shape: granularity reference + data station ──────────────
# The `model input_2.32.xlsx` deliverable is two sheets: a per-indicator model
# granularity reference (渠道 scope × 区域 granularity; blank = not selected) and
# the long-format Data Station (the model-input rows at those granularities).

_NA_TOKENS = {"", "nan", "none", "na", "<na>", "total"}


def adopted_indicators(st: ProjectState) -> dict[tuple[str, str], dict]:
    """Every ``(l4, metric)`` the model input is built from, and why it is in.

    The one derivation all three master-data surfaces read — the granularity
    reference, the Data Station and the export. They used to answer the same
    question three different ways and disagree:

    * the **response was missing from two of them**. `data_station` and
      `granularity_reference` filtered on the ledger's adopted keys, and the
      response has no ledger row — no layer rules on it, because it is what the
      drivers explain. The exported 2.32 "model input" therefore shipped without
      its dependent variable: on the drill case, 374 of 2939 rows silently gone.
    * the **granularity sheet joined in the wrong key space**, matching the factor
      row's declared ``(l4, indicator)`` against the data's ``(l4, metric)``. Those
      differ whenever a human pinned one to the other, which is exactly when a
      mapping exists at all — see :mod:`app.agents.factor_link`.

    Returns ``{(norm_l4, norm_metric): {"role", "rowId", "l1".."l4", "indicator"}}``
    where role is ``"response"`` or ``"driver"``.
    """
    from app.agents import factor_link
    from app.agents.ledger import indicator_ledger

    link = factor_link.build(st)
    out: dict[tuple[str, str], dict] = {}

    for r in indicator_ledger(st):
        if not r.adopted:
            continue
        out.setdefault(r.key, {
            "role": "driver", "rowId": link.row_for(r.l4, r.metric or r.indicator),
            "l1": r.l1, "l2": r.l2, "l3": r.l3, "l4": r.l4,
            "indicator": r.metric or r.indicator,
        })

    # The response, from the data's own role tag — the same predicate the fit,
    # 2.3's charts and `master_table` use, so the four cannot headline different Ys.
    try:
        from app.agents.dataset_cache import raw_long_df
        raw = raw_long_df(st)
        if not raw.empty:
            kpi = raw[_kpi_mask(raw, st)]
            for (l1, l2, l3, l4, metric), _g in kpi.groupby(
                    [kpi["l1"].astype("string"), kpi["l2"].astype("string"),
                     kpi["l3"].astype("string"), kpi["l4"].astype("string"),
                     kpi["metric"].astype("string")], dropna=False):
                key = (_lower(l4), _lower(metric))
                out[key] = {
                    "role": "response", "rowId": link.row_for(l4, metric),
                    "l1": _text(l1), "l2": _text(l2), "l3": _text(l3), "l4": _text(l4),
                    "indicator": _text(metric),
                }
    except Exception:  # noqa: BLE001 — no bound data yet
        pass
    return out


def _clean_set(series: pd.Series) -> list[str]:
    vals = series.astype("string").str.strip().dropna().tolist()
    return sorted({v for v in vals if v and v.lower() not in _NA_TOKENS})


def _channel_scope(channels: list[str]) -> str:
    """渠道 scope string: 全渠道 when the indicator is national/all-channel data
    (no specific channel_type in the raw rows), else the channel-type list."""
    return "全渠道" if not channels else ",".join(channels)


def _region_scope(regions: list[str]) -> str:
    """区域 granularity: National when the indicator is only national, else the
    province-group set (e.g. A,B,C,D)."""
    letters = [r for r in regions if r.lower() != "national"]
    if not letters:
        return "National"
    return ",".join(sorted(letters))


def granularity_reference(st: ProjectState) -> list[dict]:
    """The 模型颗粒度参考表 sheet: every active factor row with its 渠道 scope and
    区域 granularity (a data fact from the raw per-channel table), blank when the
    indicator was not adopted into the model."""
    from app.agents import factor_link
    from app.agents.dataset_cache import raw_long_df
    from app.agents.ledger import _norm_pair
    from app.dataeng.mapping import resolve_factor_map

    raw = raw_long_df(st)
    cov: dict[tuple[str, str], tuple[list[str], list[str]]] = {}
    if not raw.empty and {"l4", "metric", "channel_type", "province_group"} <= set(raw.columns):
        for (l4, metric), sub in raw.groupby(
                [raw["l4"].astype("string"), raw["metric"].astype("string")], dropna=False):
            cov[_norm_pair(l4, metric)] = (_clean_set(sub["channel_type"]),
                                           _clean_set(sub["province_group"]))

    adopted = adopted_indicators(st)
    link = factor_link.build(st)
    rows: list[dict] = []
    claimed: set[tuple[str, str]] = set()
    for fm in resolve_factor_map(st).rows:
        # A factor row is adopted when any DATA key that supplies it is adopted.
        # Matching `(fm.l4, fm.indicator)` against the data's keys is what left
        # PPI reading as un-adopted here while the factor-tree tab called it
        # adopted — the tree says `单品折扣 · PPI`, the data says `促销优惠 · ppi`.
        # Only keys this row actually *owns*. Two sibling factors can be pinned to
        # one data key (both `买N赠N` and `赠品小样` collect 花费 under 促销优惠);
        # `row_of_pair` picks one owner, and `factor_tree_verdicts` reports the
        # other as `notModeled`. Counting the key for both here is what made this
        # sheet claim one more adopted factor than the close-out did.
        mine = sorted(k for k in link.keys_for(fm.row_id)
                      if k in adopted and link.row_of_pair.get(k) == fm.row_id)
        claimed.update(mine)
        channels: list[str] = []
        regions: list[str] = []
        for k in mine:
            c, r = cov.get(k, ([], []))
            channels = sorted(set(channels) | set(c))
            regions = sorted(set(regions) | set(r))
        rows.append({
            "l1": fm.l1, "l2": fm.l2, "l3": fm.l3, "l4": fm.l4,
            "indicator": fm.indicator or fm.metric,
            "adopted": bool(mine),
            "role": "driver",
            # Blank channel + region = not selected into the model (2.32 convention).
            "channelScope": _channel_scope(channels) if mine else "",
            "regionScope": _region_scope(regions) if mine else "",
        })

    # Adopted keys no active factor row claims — above all the response, which is
    # not in the tree because factors explain it rather than declare it. Omitting
    # them made the sheet describe a model input it does not contain.
    for key, meta in sorted(adopted.items()):
        if key in claimed:
            continue
        channels, regions = cov.get(key, ([], []))
        rows.append({
            "l1": meta["l1"], "l2": meta["l2"], "l3": meta["l3"], "l4": meta["l4"],
            "indicator": meta["indicator"], "adopted": True, "role": meta["role"],
            "channelScope": _channel_scope(channels),
            "regionScope": _region_scope(regions),
        })
    return rows


# The Data Station columns, in the 2.32 reference order.
DATA_STATION_COLS = [
    ("task_name", "Task name"), ("brand", "品牌"), ("province_group", "省份组别"),
    ("channel_type", "渠道类型"), ("channel", "渠道"), ("year", "年"), ("month", "月"),
    ("source", "数据源"), ("l1", "数据类型Level1"), ("l2", "数据类型Level2"),
    ("l3", "数据类型Level3"), ("l4", "数据类型Level4"), ("l5", "数据类型Level5"),
    ("l6", "数据类型Level6"), ("l7", "数据类型Level7"), ("l8", "数据类型Level8"),
    ("metric_type", "METRICS类型"), ("metric", "METRICS"), ("value", "VALUE"),
]
DATA_STATION_CAP = 5000


def data_station(st: ProjectState, *, limit: int = DATA_STATION_CAP) -> dict:
    """The D.Data Station sheet: the raw per-channel long rows for the **adopted**
    indicators — the response included — at their native channel/region/time
    granularity.

    "Response included" is the whole of a bug fix: this filtered on the ledger's
    adopted keys, and the response has no ledger row, so the model-input export
    shipped without its dependent variable.
    """
    from app.agents.dataset_cache import raw_long_df
    from app.agents.ledger import _norm_pair

    raw = raw_long_df(st)
    if raw.empty:
        return {"columns": [c for _, c in DATA_STATION_COLS], "rows": [],
                "rowCount": 0, "truncated": False}
    adopted_keys = set(adopted_indicators(st))
    keys = [_norm_pair(a, b) for a, b in zip(raw["l4"].astype("string"), raw["metric"].astype("string"))]
    mask = pd.Series([k in adopted_keys for k in keys], index=raw.index)
    sub = raw[mask]
    total = len(sub)
    truncated = total > limit
    sub = sub.head(limit)
    rows = [
        [None if pd.isna(r.get(col)) else (round(float(r[col]), 2) if col == "value"
         else str(r.get(col, ""))) for col, _ in DATA_STATION_COLS]
        for _, r in sub.iterrows()
    ]
    return {"columns": [c for _, c in DATA_STATION_COLS], "rows": rows,
            "rowCount": total, "truncated": truncated}


def _kpi_metric(df: pd.DataFrame) -> str:
    """The primary KPI in a slice — prefer the Volume KPI (matches the OLS default Y)."""
    kpi_rows = df[_kpi_mask(df)]
    if kpi_rows.empty:
        return ""
    from app.agents.indicator_metadata import classify_indicator
    names = _distinct(kpi_rows, "metric")
    volume = [m for m in names if classify_indicator(m).metric_type == "kpi_volume"]
    if volume:
        return volume[0]
    mode = kpi_rows["metric"].mode()
    return str(mode.iloc[0]) if not mode.empty else ""


def _wide_frame(st: ProjectState, df: pd.DataFrame, grain: str):
    """The full (uncapped) KPI-first wide frame for a slice → (cols, wide indexed by period)."""
    from app.dataeng.validation_query import _metric_agg, _pandas_agg
    kpi_metric = _kpi_metric(df)
    keys = _period_keys(df, grain)
    frame = pd.DataFrame({
        "_k": keys,
        "_m": df["metric"].astype("string").str.strip(),
        "_v": pd.to_numeric(df["value"], errors="coerce"),
    }).dropna(subset=["_k", "_m"])
    if frame.empty:
        return [], pd.DataFrame()
    cols_data: dict[str, pd.Series] = {}
    for m, g in frame.groupby("_m"):
        cols_data[str(m)] = g.groupby("_k")["_v"].agg(_pandas_agg(_metric_agg(st, str(m))))
    wide = pd.DataFrame(cols_data).sort_index()
    cols = [c for c in wide.columns if _lower(c) == _lower(kpi_metric)]
    cols += sorted(c for c in wide.columns if _lower(c) != _lower(kpi_metric))
    return cols, wide


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet titles: ≤31 chars, none of ``[]:*?/\\``, and unique."""
    clean = "".join("-" if c in "[]:*?/\\" else c for c in name).strip() or "model"
    title = clean[:31]
    n = 2
    while title in used:
        suffix = f"~{n}"
        title = clean[:31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def model_input_sheets(st: ProjectState, grain: str = "month") -> list[tuple[str, list[str], list[list]]]:
    """One wide table per model object — the feature matrix modeling consumes.

    The Data Station is the long form and is what the 2.32 reference ships; this
    is the same rows pivoted the way a regression eats them (one row per period,
    one column per adopted indicator, response first), so the export carries the
    model input itself and not only the evidence behind it.
    """
    from app.agents.dataset_cache import model_objects
    from app.agents.model_objects import object_label, object_mask

    out: list[tuple[str, list[str], list[list]]] = []
    for obj in (model_objects(st) or []):
        # Screened against THIS object only, so a national indicator another
        # channel rejected still appears in the model that actually used it.
        try:
            df = _adopted_df(st, scope=[obj])
        except Exception:  # noqa: BLE001 — no bound data yet
            return []
        if df.empty:
            continue
        # `object_mask` is the row selection the FIT uses (`pivot._resolve_object_filter`).
        # Slicing by brand + channel_type instead looks equivalent and is not: it drops
        # every national row (blank channel) and every competitor row (a brand this
        # object does not name), which are exactly the shared rows the model is fitted
        # on. The TT sheet came out with nothing but Y for that reason.
        sub = df[object_mask(df, obj)]
        if sub.empty:
            continue
        g = grain if grain in _available_grains(sub) else (
            "month" if "month" in _available_grains(sub) else "year")
        cols, wide = _wide_frame(st, sub, g)
        if not cols or wide.empty:
            continue
        rows = [[_period_label(int(k), g)]
                + [None if pd.isna(v := wide.at[k, c]) else round(float(v), 4) for c in cols]
                for k in list(wide.index)]
        out.append((object_label(obj), ["Period"] + [str(c) for c in cols], rows))
    return out


def build_export(st: ProjectState, **_ignored) -> bytes:
    """The 2.32 ``model input`` deliverable as xlsx, uncapped:

    * ``模型颗粒度参考表`` — every factor row with its 渠道 scope × 区域 granularity,
      blank when it was not adopted;
    * ``D.Data Station`` — the adopted indicators' raw long rows, response included;
    * one sheet per model object — the same rows as the wide feature matrix.

    All three read :func:`adopted_indicators`, so the workbook cannot describe an
    indicator set the model was not built on.
    """
    from io import BytesIO
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1 — 模型颗粒度参考表
    ws1 = wb.create_sheet(title="模型颗粒度参考表")
    ws1.append(["生意因子-Level 1", "生意因子-Level 2", "生意因子-Level 3",
                "生意影响因子-Level 4", "指标选择", "角色", "渠道", "区域"])
    for r in granularity_reference(st):
        ws1.append([r["l1"], r["l2"], r["l3"], r["l4"], r["indicator"],
                    r.get("role", "driver") if r["adopted"] else "",
                    r["channelScope"], r["regionScope"]])

    # Sheet 2 — D.Data Station (uncapped)
    ws2 = wb.create_sheet(title="D.Data Station")
    ds = data_station(st, limit=10_000_000)
    ws2.append(ds["columns"])
    for row in ds["rows"]:
        ws2.append(row)

    # Sheets 3+ — the wide feature matrix, one per channel × product model.
    used = {"模型颗粒度参考表", "D.Data Station"}
    for label, columns, rows in model_input_sheets(st):
        ws = wb.create_sheet(title=_safe_sheet_title(label, used))
        ws.append(columns)
        for row in rows:
            ws.append(row)

    if not wb.sheetnames:
        wb.create_sheet(title="empty")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
