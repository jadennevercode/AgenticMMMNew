"""Data-request upload manifest — the BU-derived L3 directory.

The Business-Understanding Data Request (a-data-request) is organised as one
workbook per L3, one sheet per L4, columns = time + scope + the L4 indicators.
This module turns that contract into an upload manifest of **L3 slots**, binds
each uploaded data file to its slot, and validates per-slot coverage (are the
expected L4 sheets + indicators present in the uploaded workbook?). This is what
guarantees every upload maps back to the factor tree.
"""
from __future__ import annotations

import re
from pathlib import Path

from app.domain.models import DataRequestManifest, DataRequestSlot
from app.store.files import get_files
from app.store.state import ProjectState

_MISSING_CAP = 30


def _norm(s: object) -> str:
    """Normalise a label for matching: keep alphanumerics + CJK, drop the rest."""
    return re.sub(r"[^0-9a-z一-鿿]", "", str(s).lower())


def factor_tree_by_l3(st: ProjectState) -> dict[str, dict[str, list[str]]]:
    """{L3: {L4: [indicators]}} over accepted/baseline factor-tree rows."""
    out: dict[str, dict[str, list[str]]] = {}
    ft = st.factor_tree
    if ft is None:
        return out
    for r in ft.rows:
        if r.status not in ("baseline", "accepted"):
            continue
        l3 = r.l3 or r.l2 or r.l1 or "—"
        l4 = r.l4 or l3
        out.setdefault(l3, {}).setdefault(l4, [])
        if r.indicator and r.indicator not in out[l3][l4]:
            out[l3][l4].append(r.indicator)
    return out


def _workbook_headers(path: Path) -> dict[str, list[str]]:
    """{sheet_title: [header columns]} from an xlsx; {} if unreadable/not xlsx."""
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — never let a bad upload break the manifest
        return {}
    out: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        header: list[str] = []
        for row in ws.iter_rows(values_only=True):
            header = ["" if c is None else str(c) for c in row]
            break
        out[ws.title] = header
    wb.close()
    return out


def _match_score(sheet_norm: str, l4_norm: str) -> int:
    """Higher = better sheet↔L4 match. Exact > prefix > L4-in-sheet > sheet-in-L4."""
    if not sheet_norm:
        return 0
    if sheet_norm == l4_norm:
        return 4
    if sheet_norm.startswith(l4_norm) or l4_norm.startswith(sheet_norm):
        return 3  # handles Excel's 31-char title truncation
    if l4_norm in sheet_norm:
        return 2
    if sheet_norm in l4_norm:
        return 1  # weakest — avoids 'tv' hijacking 'otvott'
    return 0


def _best_sheet(l4_norm: str, norm_sheets: dict[str, list[str]]) -> list[str] | None:
    scored = [(_match_score(ns, l4_norm), len(ns), cols) for ns, cols in norm_sheets.items()]
    scored = [t for t in scored if t[0] > 0]
    if not scored:
        return None
    scored.sort(key=lambda t: (t[0], t[1]), reverse=True)  # best score, then longest name
    return scored[0][2]


def _score_slot(slot: DataRequestSlot, l4s: dict[str, list[str]], headers: dict[str, list[str]]) -> None:
    """Fill coverage/missing/status on `slot` from the uploaded workbook headers."""
    norm_sheets = {_norm(name): cols for name, cols in headers.items()}
    missing_l4: list[str] = []
    missing_ind: list[str] = []
    covered = 0
    for l4, inds in l4s.items():
        nl4 = _norm(l4)
        cols = _best_sheet(nl4, norm_sheets)
        if cols is None:
            missing_l4.append(l4)
            missing_ind.extend(f"{l4}·{i}" for i in (inds or []))
            continue
        ncols = [_norm(c) for c in cols if c]
        for ind in inds or []:
            ni = _norm(ind)
            if ni and any(ni == c or ni in c for c in ncols):
                covered += 1
            else:
                missing_ind.append(f"{l4}·{ind}")
    slot.covered_indicators = covered
    slot.missing_l4s = missing_l4
    slot.missing_indicators = missing_ind[:_MISSING_CAP]
    slot.status = "validated" if not missing_l4 and not missing_ind else "incomplete"


def build_manifest(st: ProjectState) -> DataRequestManifest:
    """Build the L3-slot upload manifest with per-slot coverage status."""
    by_l3 = factor_tree_by_l3(st)
    files = get_files()
    bound = {f.slot: f for f in files.list(st.project_id) if f.category == "data" and f.slot}

    prof = st.profile
    gran = prof.time_granularity if prof else "Month"
    dims = [d.name for d in prof.model_scope.dimensions] if prof else []

    # Data Engine: published indicators cover L3 slots without a slot upload — the
    # transformed asset already carries the factor path, so per-L3 coverage counts
    # indicators as first-class evidence alongside uploaded workbooks.
    def _norm(s: str) -> str:
        return "".join(str(s).lower().split())
    indicator_l3 = {}
    for ind in getattr(st, "indicators", []) or []:
        if ind.l3:
            indicator_l3.setdefault(_norm(ind.l3), []).append(ind)

    slots: list[DataRequestSlot] = []
    validated = 0
    for l3, l4s in by_l3.items():
        expected = sum(len(inds or []) for inds in l4s.values())
        slot = DataRequestSlot(l3=l3, expectedL4s=list(l4s.keys()), expectedIndicators=expected)
        rec = bound.get(l3)
        if rec is not None:
            slot.file_id = rec.id
            slot.filename = rec.filename
            slot.status = "uploaded"
            found = files.get_path(st.project_id, rec.id)
            headers = _workbook_headers(found[1]) if found else {}
            if not headers:
                slot.status = "error"
            else:
                _score_slot(slot, l4s, headers)
        if slot.status != "validated" and _norm(l3) in indicator_l3:
            covering = indicator_l3[_norm(l3)]
            slot.status = "validated"
            slot.covered_indicators = max(slot.covered_indicators, len(covering))
            if not slot.filename:
                slot.filename = f"(published indicators: {covering[0].asset_name})"
        if slot.status == "validated":
            validated += 1
        slots.append(slot)

    return DataRequestManifest(slots=slots, total=len(slots), validated=validated,
                               timeGranularity=gran, scopeDims=dims)


def manifest_satisfied(st: ProjectState) -> bool:
    """True when the data-request manifest's per-L3 coverage is met.

    Drives the 2.1 gate (``requiresManifest``): when a manifest exists (the factor
    tree produced L3 slots), every slot must be ``validated``. When there is no
    manifest to validate against (no slots), the gate degrades to the upstream
    ``requiresUpload`` file-presence check — it does not hard-block."""
    try:
        m = build_manifest(st)
    except Exception:  # noqa: BLE001 — never let manifest errors hard-block the engine
        return True
    return m.total == 0 or m.validated == m.total


# ── Export: one .xlsx workbook per L3, one sheet per L4 ──────────────────────
_INVALID_SHEET = re.compile(r"[\\/?*\[\]:]")
_INVALID_FILE = re.compile(r'[\\/:*?"<>|]+')

# Styling to mirror reference/data request template.xlsx ──────────────────────
_FONT_NAME = "Open Sans"
_FONT_SIZE = 11
_HEADER_FILL = "FFAFC7FE"  # periwinkle header band, as in the reference template
_TITLE_COL_WIDTH = 45.3


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet title: ≤31 chars, no invalid chars, unique within a workbook."""
    base = _INVALID_SHEET.sub(" ", str(name)).strip()[:31] or "Sheet"
    title, n = base, 1
    while title.lower() in used:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def _safe_file_name(name: str, used: set[str]) -> str:
    base = _INVALID_FILE.sub("_", str(name)).strip().strip(".")[:80] or "L3"
    fname, n = base, 1
    while fname.lower() in used:
        fname = f"{base}_{n}"
        n += 1
    used.add(fname.lower())
    return fname


def _granularity_label(time_granularity: str, scope_dims: list[str]) -> str:
    """'By time, by <dim>, …' — the reference's Granularity descriptor."""
    parts = ["By time", *[f"by {d}" for d in scope_dims]]
    return ", ".join(parts)


def _write_sheet(ws, time_col: str, scope_dims: list[str], indicators: list[str],
                 scope_rows: list, granularity: str, brand: str) -> None:
    """Lay out one L4 sheet to mirror the reference template: an info header block,
    a styled (periwinkle, bold, centered) column-header row, then example rows.

    Column logic is unchanged — time + model-scope dims + the L4 indicators."""
    from openpyxl.styles import Alignment, Font, PatternFill

    base_font = Font(name=_FONT_NAME, size=_FONT_SIZE)
    bold_font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True)
    header_fill = PatternFill(patternType="solid", fgColor=_HEADER_FILL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(vertical="center", wrap_text=True)

    def put(row: int, col: int, value, *, font=base_font, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align
        return cell

    # Info header block (labels left, values in column B) ──────────────────────
    put(1, 1, "Data Request", font=bold_font)
    put(2, 1, "Time Coverage: ")
    put(3, 1, "Granularity: ", align=Alignment(vertical="center"))
    put(3, 2, _granularity_label(granularity, scope_dims), align=left_wrap)
    put(4, 1, "Brand Coverage: ")
    put(4, 2, brand)
    ws.row_dimensions[3].height = 50

    # Styled column-header row (row 6, blank spacer at row 5) ───────────────────
    header_row = 6
    headers = [time_col, *scope_dims, *indicators]
    for c, label in enumerate(headers, start=1):
        put(header_row, c, label, font=bold_font, fill=header_fill, align=center)

    # Example data rows (column composition unchanged) ─────────────────────────
    r = header_row + 1
    if scope_rows:
        for sr in scope_rows:
            put(r, 1, "2023-01", align=center)
            for i in range(len(scope_dims)):
                put(r, 2 + i, sr[i] if i < len(sr) else "", align=center)
            for j in range(len(indicators)):
                put(r, 2 + len(scope_dims) + j, "", align=center)
            r += 1
    else:
        put(r, 1, "2023-01", align=center)

    ws.column_dimensions["A"].width = _TITLE_COL_WIDTH


def build_export_zip(st: ProjectState) -> bytes:
    """Build the Data Request as a ZIP of one .xlsx workbook per L3 (one sheet per
    L4; columns = time + model-scope dims + the L4 indicators, with example rows).

    Styling mirrors reference/data request template.xlsx (info header block +
    periwinkle column-header band, Open Sans 11)."""
    from io import BytesIO
    import zipfile

    import openpyxl

    by_l3 = factor_tree_by_l3(st)
    prof = st.profile
    granularity = prof.time_granularity if prof else "Month"
    time_col = f"Time ({granularity})"
    scope_dims = [d.name for d in prof.model_scope.dimensions] if prof else ["Channel"]
    scope_rows = prof.model_scope.rows[:3] if prof else []
    brand = st.meta.brand if st.meta else ""

    buf = BytesIO()
    used_files: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for l3, l4s in (by_l3 or {"Data Request": {"value": ["value"]}}).items():
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            used_sheets: set[str] = set()
            for l4, inds in (l4s or {"value": ["value"]}).items():
                ws = wb.create_sheet(title=_safe_sheet_title(l4, used_sheets))
                _write_sheet(ws, time_col, scope_dims, inds or ["value"],
                             scope_rows, granularity, brand)
            if not wb.sheetnames:  # L3 with no L4 — keep a valid workbook
                ws = wb.create_sheet(title="Sheet1")
                _write_sheet(ws, time_col, scope_dims, [], scope_rows, granularity, brand)
            wbuf = BytesIO()
            wb.save(wbuf)
            zf.writestr(f"{_safe_file_name(l3, used_files)}.xlsx", wbuf.getvalue())
    return buf.getvalue()
