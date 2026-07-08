"""Loader for the data-quality validation rubric (Data Validation_2.1).

Three sheets:
  - ``2.11数据通用校验标准``  free-text compliance principles (4 dimensions)
  - ``2.11打分规则``         the 0 / 0.5 / 1 scoring rubric per criterion
  - ``2.12数据质量评分``      per-factor scored table (L1..L4 x metric -> scores)

The public ``load_validation_rules`` returns the structured 4-dimension rubric
plus the raw compliance principles, which is what the data agent scores against.
"""
from __future__ import annotations

import openpyxl

from ._paths import VALIDATION, ref

_RUBRIC_SHEET = "2.11打分规则"
_PRINCIPLES_SHEET = "2.11数据通用校验标准"
_SCORES_SHEET = "2.12数据质量评分"

# Map the dimension banner text to a stable English key.
_DIMENSION_KEYS = {
    "Consistency": "consistency",
    "Accuracy": "accuracy",
    "Comprehensiveness": "completeness",
    "Completeness": "completeness",
    "Granularity": "granularity",
}


def _clean(v: object) -> str:
    return "" if v is None else str(v).strip()


def _match_dimension(text: str) -> str | None:
    for token, key in _DIMENSION_KEYS.items():
        if token.lower() in text.lower():
            return key
    return None


def _load_rubric() -> dict:
    """Parse the 0/0.5/1 scoring rubric into {dimension: [criteria]}."""
    wb = openpyxl.load_workbook(ref(VALIDATION), read_only=True, data_only=True)
    ws = wb[_RUBRIC_SHEET]
    rows = [[_clean(c) for c in (list(r) + ["", "", ""])[:3]]
            for r in ws.iter_rows(values_only=True)]
    wb.close()

    dimensions: dict[str, dict] = {}
    current_dim: str | None = None
    current_crit: dict | None = None

    for c0, c1, c2 in rows:
        if not any((c0, c1, c2)):
            continue
        dim = _match_dimension(c0) if not c1 and not c2 else None
        # A dimension banner: name in col0, cols1/2 empty.
        if dim and not c1 and not c2:
            current_dim = dim
            dimensions[dim] = {"label": c0, "criteria": []}
            current_crit = None
            continue
        # The "0分 / 0.5分 / 1分" header row -> skip.
        if c0.replace(" ", "") in {"0分"} and "0.5" in c1:
            continue
        if current_dim is None:
            continue
        # A criterion title: text in col0, cols1/2 empty, numbered like "1. ...".
        if c0 and not c1 and not c2:
            current_crit = {"name": c0, "score_0": "", "score_0_5": "", "score_1": ""}
            dimensions[current_dim]["criteria"].append(current_crit)
            continue
        # A scoring row: descriptions for 0 / 0.5 / 1.
        if current_crit is not None:
            current_crit["score_0"] = c0
            current_crit["score_0_5"] = c1
            current_crit["score_1"] = c2
    return dimensions


def _load_principles() -> list[str]:
    wb = openpyxl.load_workbook(ref(VALIDATION), read_only=True, data_only=True)
    ws = wb[_PRINCIPLES_SHEET]
    lines = [_clean(r[0]) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return [ln for ln in lines if ln]


def load_validation_rules() -> dict:
    """Return the 4-dimension data-quality scoring rubric as structured rules.

    Shape::

        {
          "scale": [0, 0.5, 1],
          "dimensions": {
            "consistency": {"label": ..., "criteria": [
                {"name","score_0","score_0_5","score_1"}, ...]},
            "accuracy": {...}, "completeness": {...}, "granularity": {...},
          },
          "principles": [ "<compliance principle line>", ... ],
        }
    """
    return {
        "scale": [0.0, 0.5, 1.0],
        "dimensions": _load_rubric(),
        "principles": _load_principles(),
    }
