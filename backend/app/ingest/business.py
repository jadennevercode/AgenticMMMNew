"""Loaders for business-agent reference inputs: scope, KBQs, industry knowledge.

These sheets are human-authored with banners, merged cells and Chinese text, so
loaders are best-effort: they find the real header row, then emit structured rows
while preserving the raw free-text where structure is ambiguous.
"""
from __future__ import annotations

import openpyxl

from ._paths import BIZ, INDUSTRY, KBQS, SCOPE, ref, reference_root


def _clean(v: object) -> str:
    return "" if v is None else str(v).replace("\xa0", " ").strip()


def _sheet_rows(path, sheet: str) -> list[list[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = [[_clean(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def load_scope() -> dict:
    """Parse the project scope (channel x region/platform x product)."""
    rows = _sheet_rows(ref(SCOPE), "Sheet1")
    notes = [r[0] for r in rows if r and r[0] and "渠道" not in r[0] and len(r) and not (
        len(r) >= 3 and r[1] and r[2])]
    # Find the header row that defines the scope matrix.
    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "渠道"), None)
    entries: list[dict] = []
    headers: list[str] = []
    if header_idx is not None:
        headers = [c for c in rows[header_idx] if c]
        for r in rows[header_idx + 1:]:
            if not any(r):
                continue
            if not r[0]:
                continue
            entries.append({
                "channel": r[0],
                "region_platform": r[1] if len(r) > 1 else "",
                "product": r[2] if len(r) > 2 else "",
            })
    return {
        "notes": [n for n in notes if n],
        "columns": headers,
        "entries": entries,
    }


def _kbq_path():
    """KBQs path — the current folder ships only a PNG, so fall back to the
    archived workbook which holds the real structured KBQ data."""
    try:
        return ref(KBQS)
    except FileNotFoundError:
        for cand in reference_root().glob(f"{BIZ}/[[]0.ARCHIVE/*KBQs*.xlsx"):
            return cand
        raise


def load_kbqs() -> list[dict]:
    """Parse Key Business Questions from the ``KBQs原`` sheet (structured)."""
    rows = _sheet_rows(_kbq_path(), "KBQs原")
    header_idx = next((i for i, r in enumerate(rows)
                       if "编号" in r and "关键业务问题" in " ".join(r)), None)
    out: list[dict] = []
    if header_idx is None:
        return out
    current: dict | None = None
    for r in rows[header_idx + 1:]:
        if not any(r):
            continue
        code = r[1] if len(r) > 1 else ""
        category = r[2] if len(r) > 2 else ""
        question = r[3] if len(r) > 3 else ""
        context = r[4] if len(r) > 4 else ""
        raised_by = r[5] if len(r) > 5 else ""
        raised_name = r[6] if len(r) > 6 else ""
        if code:  # new KBQ
            current = {
                "code": code,
                "category": category,
                "question": question,
                "context": [c for c in [context] if c],
                "raised_by": raised_by,
                "raised_by_name": raised_name,
            }
            out.append(current)
        elif current is not None and context:
            current["context"].append(context)
    return out


def load_industry_knowledge() -> dict:
    """Parse the industry-knowledge workbook (table of contents + frameworks)."""
    wb = openpyxl.load_workbook(ref(INDUSTRY), read_only=True, data_only=True)
    result: dict = {"sheets": wb.sheetnames, "sections": {}}
    for sn in wb.sheetnames:
        ws = wb[sn]
        lines: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            cells = [_clean(c) for c in r]
            if any(cells):
                lines.append([c for c in cells if c])
        result["sections"][sn] = lines
    wb.close()
    return result
