"""Parse a user-uploaded factor-tree workbook into FactorRow list.

The upload template mirrors the exported ``a-factor-tree`` "Factor Tree" sheet:
columns L1 · L2 · L3 · L4 · Indicator (Dimension/Source/Status are ignored on
import — provenance is re-stamped as ``source="upload"``). L1..L4 cells may be
merged (forward-filled) like a hand-authored tree. Header detection is tolerant
of English ("L1".."Indicator") and Chinese ("生意因子-level 1" / "指标") names.

Deterministic, no LLM — the AI supplementation happens later in
``business.derive_factor_tree``.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl

from app.domain.models import FactorRow

# Header aliases per logical column (lower-cased, whitespace-stripped substrings).
_ALIASES: dict[str, tuple[str, ...]] = {
    "l1": ("l1", "level 1", "level1", "生意因子-level 1", "一级"),
    "l2": ("l2", "level 2", "level2", "生意因子-level 2", "二级"),
    "l3": ("l3", "level 3", "level3", "生意因子-level 3", "三级"),
    "l4": ("l4", "level 4", "level4", "生意因子-level 4", "四级"),
    "indicator": ("indicator", "指标", "指标选择", "metric"),
}


def _clean(v: object) -> str:
    return str(v).strip() if v is not None else ""


def _match_col(header: str) -> Optional[str]:
    h = header.strip().lower()
    if not h:
        return None
    for logical, aliases in _ALIASES.items():
        if any(a in h for a in aliases):
            return logical
    return None


def _column_map(rows: list[list[str]]) -> Optional[tuple[int, dict[str, int]]]:
    """Find the header row (within the first 6 rows) and map logical→col index.

    Returns (header_row_index, {logical: col}) or None when no L1/indicator
    columns can be located."""
    for ri, row in enumerate(rows[:6]):
        mapping: dict[str, int] = {}
        for ci, cell in enumerate(row):
            logical = _match_col(cell)
            if logical and logical not in mapping:
                mapping[logical] = ci
        if "indicator" in mapping and ("l1" in mapping or "l4" in mapping):
            return ri, mapping
    return None


def _parse_sheet(ws) -> list[FactorRow]:
    grid = [[_clean(c) for c in r] for r in ws.iter_rows(values_only=True)]
    found = _column_map(grid)
    if found is None:
        return []
    header_ri, cmap = found
    carry = {k: "" for k in ("l1", "l2", "l3", "l4")}
    out: list[FactorRow] = []
    for r in grid[header_ri + 1:]:
        def cell(logical: str) -> str:
            ci = cmap.get(logical, -1)
            return r[ci] if 0 <= ci < len(r) else ""
        # Forward-fill merged L1..L4.
        for lvl in ("l1", "l2", "l3", "l4"):
            v = cell(lvl)
            if v:
                carry[lvl] = v
        indicator = cell("indicator")
        if not indicator and not carry["l4"]:
            continue  # blank spacer row
        out.append(FactorRow(
            id=f"ft-up-{len(out)}",
            l1=carry["l1"], l2=carry["l2"], l3=carry["l3"], l4=carry["l4"],
            indicator=indicator, source="upload", status="baseline"))
    return out


def parse_factor_tree_upload(paths: list[Path]) -> list[FactorRow]:
    """Parse one or more uploaded factor-tree workbooks into FactorRow list.

    Prefers a sheet named "Factor Tree" (the export sheet); otherwise scans every
    sheet and keeps the first that yields rows. Non-xlsx / unreadable files are
    skipped. Returns [] when nothing parses (caller then blocks the deliverable)."""
    rows: list[FactorRow] = []
    for path in paths:
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            continue
        sheet_order = ([wb["Factor Tree"]] if "Factor Tree" in wb.sheetnames else []) + [
            wb[n] for n in wb.sheetnames if n != "Factor Tree"]
        for ws in sheet_order:
            parsed = _parse_sheet(ws)
            if parsed:
                rows.extend(parsed)
                break
    # Re-key so ids are unique across files.
    return [r.model_copy(update={"id": f"ft-up-{i}"}) for i, r in enumerate(rows)]
