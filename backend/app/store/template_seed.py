"""Seed the built-in beverage templates from the Assets reference workbooks.

Parsed ONCE on first template-store access and persisted to data/templates/.
Source assets live in the repo's ``Assets/`` folder:
    - Default Factor Tree.xlsx        -> beverage factor-tree template
    - Interview Plan & Question.xlsx   -> beverage interview-outline template
If the assets are absent, seeding yields empty built-ins (non-fatal).
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from app.config import BACKEND_ROOT
from app.domain.models import (
    GENERAL_INDUSTRY,
    FactorTreeRow,
    InterviewQuestion,
    KnowledgeNote,
    KnowledgeTemplate,
    RuleRow,
)

ASSETS_DIR = BACKEND_ROOT.parent / "Assets"
_FACTOR_TREE_FILE = "Default Factor Tree.xlsx"
_INTERVIEW_FILE = "Interview Plan & Question.xlsx"

# Beverage built-ins are keyed to the food-bev > beverage industry node.
_BEV_L1 = "food-bev"
_BEV_L2 = "beverage"

# Built-in MMM methodology rules (S2 quality / statistical / technical review) +
# a beverage business rule. Ported from the legacy Solution-Library content so the
# seeded beverage pack ships a real, mappable rules section.
_BEV_RULES: list[RuleRow] = [
    RuleRow(id="rule-quality", category="quality", severity="block",
            name="数据质量打分规则",
            detail="一致性 / 准确性 / 完整性 / 颗粒度 四维，0 / 0.5 / 1 三态打分；"
                   "任一维度为 0 判 unusable，最终分取四维最小值。"),
    RuleRow(id="rule-stat", category="statistical", severity="warn",
            name="统计初筛规则",
            detail="CV · Pearson · VIF 组合打分：综合 ≥3 记 Good，<1.5 剔除；"
                   "高共线（VIF 高）变量考虑合并。"),
    RuleRow(id="rule-tech", category="technical", severity="warn",
            name="技术审查基准",
            detail="R² 85–95% · MAPE 5–15% · DW 1.5–2.5；越界触发红旗规则，需复核而非直接采用。"),
    RuleRow(id="rule-wholesale", category="business", severity="info",
            name="批发并入 TT 渠道规则",
            detail="批发数据不可拆时并入 TT 渠道，避免稀释 AFH 投资有效性。"),
]

# Built-in industry know-how notes (beverage-specific narrative knowledge).
_BEV_KNOWLEDGE: list[KnowledgeNote] = [
    KnowledgeNote(id="kn-throat", title="Share of Throat（跨品类挤压）",
                  body="现调饮料 / 外卖即饮可作为饮料品类的外部竞争因子；外卖大战等结构性"
                       "突变需作为单独外部变量纳入，否则会被误吸收进媒体效应。",
                  tags=["外部因素", "竞争"]),
    KnowledgeNote(id="kn-cooler", title="冰柜 × 陈列资产共线",
                  body="终端冷柜与陈列资产高度共线时应合并为单一渠道执行变量，避免 VIF 膨胀。",
                  tags=["渠道执行", "共线"]),
]

# Built-in general (cross-industry) knowledge — method & style, used for grounding.
_GENERAL_KNOWLEDGE: list[KnowledgeNote] = [
    KnowledgeNote(id="gk-narrative", title="报告叙事风格",
                  body="每个结论同时给出量级与占比；ROI 结论必须附「不可线性外推」的口径说明。",
                  tags=["报告", "叙事"]),
    KnowledgeNote(id="gk-chart", title="Chart book 骨架",
                  body="Part 0 基本面 → Part 1 全国 → Part 2 区域，逐层下钻，保持同一指标口径。",
                  tags=["报告", "结构"]),
    KnowledgeNote(id="gk-interview", title="分层访谈框架",
                  body="Layer 1/2/3（高层 / 管理层 / 执行层）× 三段式问题（现状 / 期望 / 数据可得性）。",
                  tags=["访谈", "方法"]),
]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _clean(v: object) -> str:
    return "" if v is None else str(v).strip()


def _parse_factor_tree(path: Path) -> list[FactorTreeRow]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows: list[FactorTreeRow] = []
    carry = ["", "", "", ""]  # forward-fill L1..L4 (merged cells)
    for i, raw in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header banner
        cells = [_clean(c) for c in (list(raw) + [""] * 5)[:5]]
        l1, l2, l3, l4, indicator = cells
        for idx, val in enumerate((l1, l2, l3, l4)):
            if val:
                carry[idx] = val
                for j in range(idx + 1, 4):
                    carry[j] = ""
        if not any(carry) and not indicator:
            continue
        rows.append(FactorTreeRow(
            l1=carry[0], l2=carry[1], l3=carry[2], l4=carry[3], indicator=indicator))
    wb.close()
    return rows


# Map interview workbook sheet titles to the 4-category model.
def _category_for(sheet_title: str) -> str | None:
    title = sheet_title.lower()
    if "leadership" in title:
        return "Leadership"
    if title.startswith("management"):
        return "Management"
    if title.startswith("operation"):
        return "Operation"
    return None


def _parse_interview(path: Path) -> list[InterviewQuestion]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    questions: list[InterviewQuestion] = []
    for ws in wb.worksheets:
        category = _category_for(ws.title)
        if category is None:
            continue
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        in_questions = False
        pending: str | None = None
        for r in rows:
            cells = [_clean(c) for c in r]
            joined = " ".join(cells)
            if not in_questions:
                if "Questions" in joined and "Objective" not in joined:
                    in_questions = True
                continue
            # A numbered row (number in col idx 1) holds the Chinese text in col 2.
            text = next((c for c in cells[1:] if len(c) > 8), "")
            has_number = any(c.isdigit() for c in cells[1:3])
            if has_number and text:
                if pending:
                    questions.append(InterviewQuestion(
                        category=category, role=ws.title, question=pending[:400]))
                pending = text
            elif text and pending is not None and any(ch.isascii() and ch.isalpha() for ch in text):
                # English continuation row — prefer the English phrasing.
                pending = text
        if pending:
            questions.append(InterviewQuestion(
                category=category, role=ws.title, question=pending[:400]))
    wb.close()
    return questions


def build_builtin_templates() -> list[KnowledgeTemplate]:
    """Parse the Assets workbooks into the seed templates (best-effort)."""
    templates: list[KnowledgeTemplate] = []
    now = _now_iso()

    ft_path = ASSETS_DIR / _FACTOR_TREE_FILE
    factor_rows: list[FactorTreeRow] = []
    if ft_path.exists():
        try:
            factor_rows = _parse_factor_tree(ft_path)
        except Exception:  # noqa: BLE001
            factor_rows = []
    templates.append(KnowledgeTemplate(
        id="tpl-bev-factor-tree", kind="factor_tree",
        name="Beverage Factor Tree (default)", industryL1=_BEV_L1, industryL2=_BEV_L2,
        version=1, builtin=True, factorRows=factor_rows, updatedAt=now,
    ))

    iv_path = ASSETS_DIR / _INTERVIEW_FILE
    questions: list[InterviewQuestion] = []
    if iv_path.exists():
        try:
            questions = _parse_interview(iv_path)
        except Exception:  # noqa: BLE001
            questions = []
    templates.append(KnowledgeTemplate(
        id="tpl-bev-interview", kind="interview",
        name="Beverage Interview Outline (default)", industryL1=_BEV_L1, industryL2=_BEV_L2,
        version=1, builtin=True, interviewQuestions=questions, updatedAt=now,
    ))

    templates.append(KnowledgeTemplate(
        id="tpl-bev-rules", kind="rules",
        name="Beverage Validation & Business Rules", industryL1=_BEV_L1, industryL2=_BEV_L2,
        version=1, builtin=True, ruleRows=_BEV_RULES, updatedAt=now,
    ))

    templates.append(KnowledgeTemplate(
        id="tpl-bev-knowledge", kind="industry_knowledge",
        name="Beverage Industry Knowledge", industryL1=_BEV_L1, industryL2=_BEV_L2,
        version=1, builtin=True, knowledgeNotes=_BEV_KNOWLEDGE, updatedAt=now,
    ))

    templates.append(KnowledgeTemplate(
        id="tpl-general-knowledge", kind="general_knowledge",
        name="General MMM Knowledge", industryL1=GENERAL_INDUSTRY, industryL2=None,
        version=1, builtin=True, knowledgeNotes=_GENERAL_KNOWLEDGE, updatedAt=now,
    ))
    return templates
