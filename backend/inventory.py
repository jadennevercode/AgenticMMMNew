"""Phase 0 data inventory: parse the real Danone reference files to discover
the actual modeling dataset shape, sheets, columns, and row counts.
Run: .venv/bin/python inventory.py
"""
import os
import sys
from pathlib import Path

import openpyxl

REF = Path("../reference").resolve()


def inspect_xlsx(path: Path, max_sheets=40, max_cols=30, sample_rows=3):
    print(f"\n{'='*90}\nFILE: {path.relative_to(REF)}  ({path.stat().st_size/1e6:.1f} MB)")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        print(f"  !! cannot open: {e}")
        return
    print(f"  sheets ({len(wb.sheetnames)}): {wb.sheetnames[:max_sheets]}")
    for sn in wb.sheetnames[:max_sheets]:
        ws = wb[sn]
        try:
            dims = f"{ws.max_row} rows x {ws.max_column} cols"
        except Exception:  # noqa: BLE001
            dims = "?"
        # header row
        header = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header = [str(c) for c in row[:max_cols] if c is not None]
        print(f"    - [{sn}] {dims} | header: {header}")
    wb.close()


def main():
    targets = []
    for root, _dirs, files in os.walk(REF):
        for f in files:
            if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$"):
                targets.append(Path(root) / f)
    targets.sort(key=lambda p: p.stat().st_size, reverse=True)
    print(f"Found {len(targets)} Excel files under {REF}")
    only = sys.argv[1] if len(sys.argv) > 1 else None
    for t in targets:
        if only and only not in str(t):
            continue
        inspect_xlsx(t)


if __name__ == "__main__":
    main()
